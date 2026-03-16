"""
Telegram-бот для учёта товаров маленького магазина.
Библиотеки: python-telegram-bot v20+, pandas, openpyxl
Запуск: python bot.py
"""

import os
import logging
from datetime import datetime, timedelta
from functools import wraps

import pandas as pd
from openpyxl import load_workbook, Workbook
from telegram import Update
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    ConversationHandler,
    MessageHandler,
    filters,
    ContextTypes,
)

# ──────────────────────────────────────────────────────────────────────────────
# НАСТРОЙКИ
# ──────────────────────────────────────────────────────────────────────────────
BOT_TOKEN  = "8720594664:AAHWu41HWPk3K6NTU-jatYuEV3sVaekghhw"   # ← вставьте токен бота
OWNER_ID   = 1249395828               # ← вставьте ваш Telegram user_id
EXCEL_FILE = "shop.xlsx"

# Доля прибыли
SHOP_SHARE   = 0.70   # 70% магазину
SELLER_SHARE = 0.30   # 30% продавцу

logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# СХЕМЫ ЛИСТОВ
# ──────────────────────────────────────────────────────────────────────────────
SCHEMA = {
    "products":      ["product_id", "name", "buy_price", "sell_price"],
    "inventory":     ["product_id", "name", "warehouse_total",
                      "moved_to_showcase", "warehouse_left"],
    "showcase":      ["product_id", "name", "showcase_total", "sold", "showcase_left"],
    "restock":       ["date", "product_id", "name", "quantity",
                      "buy_price", "total_cost"],
    "sales":         ["date", "product_id", "name", "quantity",
                      "sell_price", "revenue", "profit"],
    "weekly_reports": ["date", "period_start", "period_end",
                       "revenue", "cash", "cashless",
                       "profit", "shop_profit", "seller_profit"],
}

NUMERIC_COLS = {
    "product_id", "buy_price", "sell_price",
    "warehouse_total", "moved_to_showcase", "warehouse_left",
    "showcase_total", "sold", "showcase_left",
    "quantity", "total_cost", "revenue", "profit",
    "cash", "cashless", "shop_profit", "seller_profit",
}

# Состояния ConversationHandler
ASK_CASH = 1

# ──────────────────────────────────────────────────────────────────────────────
# EXCEL — вспомогательные функции
# ──────────────────────────────────────────────────────────────────────────────

def init_excel():
    """Создаёт Excel-файл с нужными листами, если его ещё нет."""
    if os.path.exists(EXCEL_FILE):
        # Добавляем лист weekly_reports если его нет (для старых файлов)
        wb = load_workbook(EXCEL_FILE)
        if "weekly_reports" not in wb.sheetnames:
            ws = wb.create_sheet("weekly_reports")
            ws.append(SCHEMA["weekly_reports"])
            wb.save(EXCEL_FILE)
        return
    wb = Workbook()
    wb.remove(wb.active)
    for sheet, cols in SCHEMA.items():
        ws = wb.create_sheet(sheet)
        ws.append(cols)
    wb.save(EXCEL_FILE)
    logger.info("Excel-файл создан: %s", EXCEL_FILE)


def read_sheet(sheet: str) -> pd.DataFrame:
    df = pd.read_excel(EXCEL_FILE, sheet_name=sheet)
    cols = SCHEMA[sheet]
    for c in cols:
        if c not in df.columns:
            df[c] = 0.0 if c in NUMERIC_COLS else ""
    df = df[cols].copy()
    for c in cols:
        if c in NUMERIC_COLS:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    return df


def write_sheet(sheet: str, df: pd.DataFrame):
    df = df[SCHEMA[sheet]].copy()
    wb = load_workbook(EXCEL_FILE)
    if sheet in wb.sheetnames:
        del wb[sheet]
    ws = wb.create_sheet(sheet)
    ws.append(SCHEMA[sheet])
    for row in df.itertuples(index=False):
        ws.append([None if (isinstance(v, float) and pd.isna(v)) else v
                   for v in row])
    wb.save(EXCEL_FILE)


def next_id(df: pd.DataFrame) -> int:
    if df.empty:
        return 1
    ids = df["product_id"].dropna()
    return int(ids.max()) + 1 if len(ids) else 1


def find_product(name: str):
    df = read_sheet("products")
    mask = df["name"].astype(str).str.lower() == name.lower()
    if not mask.any():
        return None, None
    row = df[mask].iloc[0]
    return int(row["product_id"]), row.to_dict()


def pid_mask(df: pd.DataFrame, pid: int) -> pd.Series:
    return df["product_id"] == float(pid)


# ──────────────────────────────────────────────────────────────────────────────
# АВТОРИЗАЦИЯ
# ──────────────────────────────────────────────────────────────────────────────

def owner_only(func):
    @wraps(func)
    async def wrapper(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
        if update.effective_user.id != OWNER_ID:
            await update.message.reply_text("⛔ Доступ запрещён.")
            return
        return await func(update, ctx)
    return wrapper


# ──────────────────────────────────────────────────────────────────────────────
# /start
# ──────────────────────────────────────────────────────────────────────────────

@owner_only
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = (
        "👋 *Бот учёта магазина*\n\n"
        "*Товары:*\n"
        "/add\\_product `name buy sell` — добавить товар\n"
        "/edit\\_product `name field value` — изменить товар\n"
        "  поля: `name` `buy\\_price` `sell\\_price`\n"
        "/products — список товаров\n\n"
        "*Склад и витрина:*\n"
        "/restock `name qty` — пополнить склад\n"
        "/to\\_showcase `name qty` — перенести на витрину\n"
        "/sell `name qty` — продать\n"
        "/inventory — остатки склада\n"
        "/showcase — остатки витрины\n\n"
        "*Отчёты:*\n"
        "/report\\_day — отчёт за сегодня\n"
        "/report\\_week — недельный отчёт (с вводом налички)\n"
    )
    await update.message.reply_text(text, parse_mode="Markdown")


# ──────────────────────────────────────────────────────────────────────────────
# 1. Добавить товар
# ──────────────────────────────────────────────────────────────────────────────

@owner_only
async def cmd_add_product(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    args = ctx.args
    if len(args) < 3:
        await update.message.reply_text(
            "❌ Использование: /add_product <name> <buy_price> <sell_price>"
        )
        return
    try:
        name       = args[0]
        buy_price  = float(args[1])
        sell_price = float(args[2])
    except ValueError:
        await update.message.reply_text("❌ Цены должны быть числами.")
        return

    pid, _ = find_product(name)
    if pid is not None:
        await update.message.reply_text(f"⚠️ Товар «{name}» уже существует (ID {pid}).")
        return

    df_p   = read_sheet("products")
    new_id = next_id(df_p)

    df_p = pd.concat([df_p, pd.DataFrame([{
        "product_id": float(new_id), "name": name,
        "buy_price": buy_price, "sell_price": sell_price,
    }])], ignore_index=True)
    write_sheet("products", df_p)

    df_i = read_sheet("inventory")
    df_i = pd.concat([df_i, pd.DataFrame([{
        "product_id": float(new_id), "name": name,
        "warehouse_total": 0.0, "moved_to_showcase": 0.0, "warehouse_left": 0.0,
    }])], ignore_index=True)
    write_sheet("inventory", df_i)

    df_s = read_sheet("showcase")
    df_s = pd.concat([df_s, pd.DataFrame([{
        "product_id": float(new_id), "name": name,
        "showcase_total": 0.0, "sold": 0.0, "showcase_left": 0.0,
    }])], ignore_index=True)
    write_sheet("showcase", df_s)

    await update.message.reply_text(
        f"✅ Товар «{name}» добавлен (ID {new_id}).\n"
        f"   Закупочная цена: {buy_price}  |  Цена продажи: {sell_price}"
    )


# ──────────────────────────────────────────────────────────────────────────────
# 1б. Редактировать товар
# ──────────────────────────────────────────────────────────────────────────────

@owner_only
async def cmd_edit_product(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    /edit_product <name> <field> <value>
    field: name | buy_price | sell_price
    """
    args = ctx.args
    if len(args) < 3:
        await update.message.reply_text(
            "❌ Использование: /edit_product <name> <field> <value>\n"
            "   Поля: name, buy_price, sell_price"
        )
        return

    old_name = args[0]
    field    = args[1].lower()
    value    = args[2]

    allowed_fields = {"name", "buy_price", "sell_price"}
    if field not in allowed_fields:
        await update.message.reply_text(
            f"❌ Недопустимое поле «{field}».\n"
            f"   Допустимые: name, buy_price, sell_price"
        )
        return

    pid, prod = find_product(old_name)
    if pid is None:
        await update.message.reply_text(f"❌ Товар «{old_name}» не найден.")
        return

    # Проверяем тип значения
    if field in ("buy_price", "sell_price"):
        try:
            value = float(value)
        except ValueError:
            await update.message.reply_text("❌ Цена должна быть числом.")
            return
    else:
        # field == "name" — проверим, не занято ли новое имя
        if value.lower() != old_name.lower():
            existing_pid, _ = find_product(value)
            if existing_pid is not None:
                await update.message.reply_text(
                    f"⚠️ Товар с именем «{value}» уже существует."
                )
                return

    # ── products ──
    df_p = read_sheet("products")
    mask = pid_mask(df_p, pid)
    df_p.loc[mask, field] = value
    write_sheet("products", df_p)

    # ── inventory и showcase — обновляем имя если менялось ──
    if field == "name":
        df_i = read_sheet("inventory")
        df_i.loc[pid_mask(df_i, pid), "name"] = value
        write_sheet("inventory", df_i)

        df_s = read_sheet("showcase")
        df_s.loc[pid_mask(df_s, pid), "name"] = value
        write_sheet("showcase", df_s)

    new_prod = read_sheet("products")
    row = new_prod[pid_mask(new_prod, pid)].iloc[0]
    await update.message.reply_text(
        f"✅ Товар обновлён:\n"
        f"   ID: {pid}\n"
        f"   Название: {row['name']}\n"
        f"   Закупочная цена: {row['buy_price']}\n"
        f"   Цена продажи: {row['sell_price']}"
    )


# ──────────────────────────────────────────────────────────────────────────────
# 1в. Список товаров
# ──────────────────────────────────────────────────────────────────────────────

@owner_only
async def cmd_products(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    df = read_sheet("products")
    if df.empty:
        await update.message.reply_text("Товаров нет.")
        return
    lines = ["📋 *Товары:*\n"]
    for _, r in df.iterrows():
        lines.append(
            f"• *{r['name']}* (ID {int(r['product_id'])})\n"
            f"  Закупка: {r['buy_price']}  |  Продажа: {r['sell_price']}\n"
        )
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


# ──────────────────────────────────────────────────────────────────────────────
# 2. Пополнение склада
# ──────────────────────────────────────────────────────────────────────────────

@owner_only
async def cmd_restock(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    args = ctx.args
    if len(args) < 2:
        await update.message.reply_text("❌ Использование: /restock <name> <quantity>")
        return
    try:
        name = args[0]
        qty  = int(args[1])
        if qty <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❌ Количество должно быть положительным целым числом.")
        return

    pid, prod = find_product(name)
    if pid is None:
        await update.message.reply_text(f"❌ Товар «{name}» не найден.")
        return

    buy_price  = float(prod["buy_price"])
    total_cost = round(qty * buy_price, 2)

    df_r = read_sheet("restock")
    df_r = pd.concat([df_r, pd.DataFrame([{
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "product_id": float(pid), "name": name,
        "quantity": float(qty), "buy_price": buy_price,
        "total_cost": total_cost,
    }])], ignore_index=True)
    write_sheet("restock", df_r)

    df_i = read_sheet("inventory")
    mask = pid_mask(df_i, pid)
    df_i.loc[mask, "warehouse_total"] = df_i.loc[mask, "warehouse_total"] + qty
    df_i.loc[mask, "warehouse_left"]  = df_i.loc[mask, "warehouse_left"]  + qty
    write_sheet("inventory", df_i)

    await update.message.reply_text(
        f"✅ Склад пополнен: «{name}» +{qty} шт.\n"
        f"   Стоимость закупки: {total_cost}"
    )


# ──────────────────────────────────────────────────────────────────────────────
# 3. Перенос на витрину
# ──────────────────────────────────────────────────────────────────────────────

@owner_only
async def cmd_to_showcase(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    args = ctx.args
    if len(args) < 2:
        await update.message.reply_text("❌ Использование: /to_showcase <name> <quantity>")
        return
    try:
        name = args[0]
        qty  = int(args[1])
        if qty <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❌ Количество должно быть положительным целым числом.")
        return

    pid, _ = find_product(name)
    if pid is None:
        await update.message.reply_text(f"❌ Товар «{name}» не найден.")
        return

    df_i   = read_sheet("inventory")
    mask_i = pid_mask(df_i, pid)
    left   = float(df_i.loc[mask_i, "warehouse_left"].values[0])

    if left < qty:
        await update.message.reply_text(
            f"❌ На складе недостаточно товара.\n"
            f"   Доступно: {int(left)} шт., запрошено: {qty} шт."
        )
        return

    df_i.loc[mask_i, "warehouse_left"]    = left - qty
    df_i.loc[mask_i, "moved_to_showcase"] = df_i.loc[mask_i, "moved_to_showcase"] + qty
    write_sheet("inventory", df_i)

    df_s   = read_sheet("showcase")
    mask_s = pid_mask(df_s, pid)
    df_s.loc[mask_s, "showcase_total"] = df_s.loc[mask_s, "showcase_total"] + qty
    df_s.loc[mask_s, "showcase_left"]  = df_s.loc[mask_s, "showcase_left"]  + qty
    write_sheet("showcase", df_s)

    await update.message.reply_text(
        f"✅ Перенесено на витрину: «{name}» {qty} шт."
    )


# ──────────────────────────────────────────────────────────────────────────────
# 4. Продажа
# ──────────────────────────────────────────────────────────────────────────────

@owner_only
async def cmd_sell(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    args = ctx.args
    if len(args) < 2:
        await update.message.reply_text("❌ Использование: /sell <name> <quantity>")
        return
    try:
        name = args[0]
        qty  = int(args[1])
        if qty <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❌ Количество должно быть положительным целым числом.")
        return

    pid, prod = find_product(name)
    if pid is None:
        await update.message.reply_text(f"❌ Товар «{name}» не найден.")
        return

    df_s = read_sheet("showcase")
    mask = pid_mask(df_s, pid)
    left = float(df_s.loc[mask, "showcase_left"].values[0])

    if left < qty:
        await update.message.reply_text(
            f"❌ На витрине недостаточно товара.\n"
            f"   Доступно: {int(left)} шт., запрошено: {qty} шт."
        )
        return

    sell_price = float(prod["sell_price"])
    buy_price  = float(prod["buy_price"])
    revenue    = round(qty * sell_price, 2)
    profit     = round(qty * (sell_price - buy_price), 2)

    df_s.loc[mask, "showcase_left"] = left - qty
    df_s.loc[mask, "sold"]          = df_s.loc[mask, "sold"] + qty
    write_sheet("showcase", df_s)

    df_sa = read_sheet("sales")
    df_sa = pd.concat([df_sa, pd.DataFrame([{
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "product_id": float(pid), "name": name,
        "quantity": float(qty), "sell_price": sell_price,
        "revenue": revenue, "profit": profit,
    }])], ignore_index=True)
    write_sheet("sales", df_sa)

    shop_cut   = round(profit * SHOP_SHARE, 2)
    seller_cut = round(profit * SELLER_SHARE, 2)

    await update.message.reply_text(
        f"✅ Продажа: «{name}» {qty} шт.\n"
        f"   Выручка: {revenue}  |  Прибыль: {profit}\n"
        f"   💼 Магазин (70%): {shop_cut}  |  🧑 Продавец (30%): {seller_cut}"
    )


# ──────────────────────────────────────────────────────────────────────────────
# 5. Склад
# ──────────────────────────────────────────────────────────────────────────────

@owner_only
async def cmd_inventory(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    df = read_sheet("inventory")
    if df.empty:
        await update.message.reply_text("Склад пуст.")
        return
    lines = ["📦 *Склад:*\n"]
    for _, r in df.iterrows():
        lines.append(
            f"• *{r['name']}*\n"
            f"  Всего поступило: {int(r['warehouse_total'])}\n"
            f"  Вынесено на витрину: {int(r['moved_to_showcase'])}\n"
            f"  Осталось на складе: {int(r['warehouse_left'])}\n"
        )
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


# ──────────────────────────────────────────────────────────────────────────────
# 6. Витрина
# ──────────────────────────────────────────────────────────────────────────────

@owner_only
async def cmd_showcase_view(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    df = read_sheet("showcase")
    if df.empty:
        await update.message.reply_text("Витрина пуста.")
        return
    lines = ["🛍 *Витрина:*\n"]
    for _, r in df.iterrows():
        lines.append(
            f"• *{r['name']}*\n"
            f"  Всего на витрине: {int(r['showcase_total'])}\n"
            f"  Продано: {int(r['sold'])}\n"
            f"  Осталось: {int(r['showcase_left'])}\n"
        )
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


# ──────────────────────────────────────────────────────────────────────────────
# 7. Дневной отчёт
# ──────────────────────────────────────────────────────────────────────────────

def _calc_report(days: int):
    """Возвращает dict с данными отчёта или None если продаж нет."""
    df = read_sheet("sales")
    if df.empty:
        return None

    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    since = datetime.now() - timedelta(days=days)
    df_f  = df[df["date"] >= since]

    if df_f.empty:
        return None

    total_revenue = round(df_f["revenue"].sum(), 2)
    total_profit  = round(df_f["profit"].sum(), 2)
    shop_profit   = round(total_profit * SHOP_SHARE, 2)
    seller_profit = round(total_profit * SELLER_SHARE, 2)

    items = (
        df_f.groupby("name")["quantity"]
        .sum()
        .reset_index()
        .rename(columns={"quantity": "qty"})
    )

    return {
        "revenue":       total_revenue,
        "profit":        total_profit,
        "shop_profit":   shop_profit,
        "seller_profit": seller_profit,
        "items":         items,
        "df":            df_f,
    }


def _format_report(data: dict, title: str, cash: float = None) -> str:
    lines = [f"📊 *{title}*\n"]
    lines.append(f"Выручка: *{data['revenue']}*")

    if cash is not None:
        cashless = round(data["revenue"] - cash, 2)
        lines.append(f"  💵 Наличные: *{cash}*")
        lines.append(f"  💳 Безнал: *{cashless}*")

    lines.append(f"\nПрибыль: *{data['profit']}*")
    lines.append(f"  💼 Магазин (70%): *{data['shop_profit']}*")
    lines.append(f"  🧑 Продавец (30%): *{data['seller_profit']}*")

    if cash is not None:
        cashless = round(data["revenue"] - cash, 2)
        # Доли от безнала
        cashless_shop   = round(cashless * SHOP_SHARE, 2)
        cashless_seller = round(cashless * SELLER_SHARE, 2)
        cash_shop       = round(cash * SHOP_SHARE, 2)
        cash_seller     = round(cash * SELLER_SHARE, 2)
        lines.append(f"\n📊 *Разбивка по типу оплаты:*")
        lines.append(f"  Наличные → магазин: *{cash_shop}*  |  продавец: *{cash_seller}*")
        lines.append(f"  Безнал   → магазин: *{cashless_shop}*  |  продавец: *{cashless_seller}*")

    lines.append("\n*Продано:*")
    for _, r in data["items"].iterrows():
        lines.append(f"  • {r['name']}: {int(r['qty'])} шт.")

    return "\n".join(lines)


@owner_only
async def cmd_report_day(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    data = _calc_report(1)
    if data is None:
        await update.message.reply_text("За сегодня продаж нет.")
        return
    await update.message.reply_text(
        _format_report(data, "Отчёт за сегодня"),
        parse_mode="Markdown",
    )


# ──────────────────────────────────────────────────────────────────────────────
# 8. Недельный отчёт — ConversationHandler (запрашиваем сумму наличных)
# ──────────────────────────────────────────────────────────────────────────────

@owner_only
async def cmd_report_week_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Шаг 1: проверяем что есть данные, спрашиваем наличные."""
    data = _calc_report(7)
    if data is None:
        await update.message.reply_text("За последние 7 дней продаж нет.")
        return ConversationHandler.END

    # Сохраняем данные в user_data для следующего шага
    ctx.user_data["week_report_data"] = data

    await update.message.reply_text(
        f"💰 Введите сумму *наличных* за неделю:\n"
        f"(Выручка за период: *{data['revenue']}*)\n\n"
        f"Или отправьте /cancel для отмены.",
        parse_mode="Markdown",
    )
    return ASK_CASH


async def cmd_report_week_cash(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Шаг 2: получаем наличные, формируем и сохраняем отчёт."""
    text = update.message.text.strip()
    try:
        cash = float(text.replace(",", "."))
        if cash < 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text(
            "❌ Введите корректную сумму числом (≥ 0). Попробуйте ещё раз или /cancel."
        )
        return ASK_CASH

    data = ctx.user_data.get("week_report_data")
    if data is None:
        await update.message.reply_text("Что-то пошло не так, начните снова.")
        return ConversationHandler.END

    if cash > data["revenue"]:
        await update.message.reply_text(
            f"❌ Сумма наличных ({cash}) превышает выручку ({data['revenue']}).\n"
            "Введите корректное значение или /cancel."
        )
        return ASK_CASH

    cash      = round(cash, 2)
    cashless  = round(data["revenue"] - cash, 2)
    now       = datetime.now()
    week_ago  = now - timedelta(days=7)

    # Записываем в лист weekly_reports
    df_wr = read_sheet("weekly_reports")
    df_wr = pd.concat([df_wr, pd.DataFrame([{
        "date":          now.strftime("%Y-%m-%d %H:%M"),
        "period_start":  week_ago.strftime("%Y-%m-%d"),
        "period_end":    now.strftime("%Y-%m-%d"),
        "revenue":       data["revenue"],
        "cash":          cash,
        "cashless":      cashless,
        "profit":        data["profit"],
        "shop_profit":   data["shop_profit"],
        "seller_profit": data["seller_profit"],
    }])], ignore_index=True)
    write_sheet("weekly_reports", df_wr)

    await update.message.reply_text(
        _format_report(data, "Недельный отчёт", cash=cash),
        parse_mode="Markdown",
    )

    ctx.user_data.pop("week_report_data", None)
    return ConversationHandler.END


async def cmd_cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.pop("week_report_data", None)
    await update.message.reply_text("❌ Отчёт отменён.")
    return ConversationHandler.END


# ──────────────────────────────────────────────────────────────────────────────
# ЗАПУСК
# ──────────────────────────────────────────────────────────────────────────────

def main():
    init_excel()

    app = ApplicationBuilder().token(BOT_TOKEN).build()

    # Недельный отчёт через ConversationHandler
    week_report_conv = ConversationHandler(
        entry_points=[CommandHandler("report_week", cmd_report_week_start)],
        states={
            ASK_CASH: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, cmd_report_week_cash),
            ],
        },
        fallbacks=[CommandHandler("cancel", cmd_cancel)],
    )

    app.add_handler(CommandHandler("start",        cmd_start))
    app.add_handler(CommandHandler("add_product",  cmd_add_product))
    app.add_handler(CommandHandler("edit_product", cmd_edit_product))
    app.add_handler(CommandHandler("products",     cmd_products))
    app.add_handler(CommandHandler("restock",      cmd_restock))
    app.add_handler(CommandHandler("to_showcase",  cmd_to_showcase))
    app.add_handler(CommandHandler("sell",         cmd_sell))
    app.add_handler(CommandHandler("inventory",    cmd_inventory))
    app.add_handler(CommandHandler("showcase",     cmd_showcase_view))
    app.add_handler(CommandHandler("report_day",   cmd_report_day))
    app.add_handler(week_report_conv)

    logger.info("Бот запущен. Ожидаем команды…")
    app.run_polling()


if __name__ == "__main__":
    main()